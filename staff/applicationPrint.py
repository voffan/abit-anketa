from openpyxl import load_workbook
from openpyxl.styles import Border, Side

import datetime


def ApplicationPrint(application,education,application_profile,exam,person,passp,snils):
	""" DANGER KEEP AWAY """

	source = "anketa\\static\\anketa\\Anketa SVFU.xlsx"
	wb = load_workbook(source)
	ws = wb.active
	#####################бюджет что-ли или нет
	budget = ''
	withfee = ''
	if application.budget:
		budget = 'V'
	ws.cell(row=9, column=11).value = budget 
	if application.withfee:
		withfee = 'V'
	ws.cell(row=9, column=26).value = withfee

	if education.doc.docType.value == 'Диплом':
		ws.cell(row=10,column=20).value = 'V'
	if education.doc.docType.value == 'Аттестат':
		ws.cell(row=10,column=9).value = 'V'				
	
	#####################институт факультет			
	for i in range(len(application.department.name)):
		ws.cell(row=11,column=i+6).value = application.department.name[i]
	#####################нфправление специальность
	for i in range(len(application_profile[0].profile.profile.edu_prog.name)):
		j=12
		r=0
		if i >= 24:
			j=13
			r=23
		ws.cell(row=j,column=i+3-r).value = application_profile[0].profile.profile.edu_prog.name[i]	
	#####################профиль
	printform1=''
	printform2=''
	printform3=''
	j=14
	r=0
	for item in application_profile:
		for i in range(len(item.profile.profile.name)):
			if i!=0:						
				if i%24==0:
					j+=1
					r+=23
			ws.cell(row=j,column=i+4-r).value = item.profile.profile.name[i]
		j+=1
	#####################форма тутже
		if item.profile.eduform=='О':
			printform1 = 'V'
		if item.profile.eduform=='З':
			printform2 = 'V'
		if item.profile.eduform=='ОЗ':
			printform3 = 'V' 
		ws.cell(row=20,column=4).value = printform1
		ws.cell(row=20,column=7).value = printform2
		ws.cell(row=20,column=10).value = printform3
	#####################экзамены ЕГЭ
	printSpecial=''
	j=23
	for item in exam:
		if j > 26:
			break
		if item.exam_examType.value=='ЕГЭ':
			ws.cell(row=j,column=4).value = item.exam_subjects.value
			ws.cell(row=j,column=17).value = item.year
			ws.cell(row=j,column=23).value = item.points
			j+=1
		if item.special:
			printSpecial='V'
	#####################экзамены вступительные испытания
	j=30
	for item in exam:
		if j > 33:
			break
		if item.exam_examType.value=='Вступительный':
			ws.cell(row=j,column=4).value = item.exam_subjects.value
			ws.cell(row=j,column=17).value = item.exam_examType.value					
			j+=1
		if item.special==True:
			printSpecial='V'
	ws.cell(row=35,column=26).value = printSpecial
	#####################подпись лол
	
	#####################Личные данные
	for i in (range(len(person.sname))):
		if i<24:
			ws.cell(row=38,column=i+3).value = person.sname[i]#фамилия
	for i in (range(len(person.fname))):
		if i<24:
			ws.cell(row=39,column=i+3).value = person.fname[i]#имя
	for i in (range(len(person.mname))):
		if i<24:
			ws.cell(row=40,column=i+3).value = person.mname[i]#отчество
	if person.sex=="М":
		ws.cell(row=41, column=4).value = 'V'#пол
	else:
		ws.cell(row=41, column=6).value = 'V'#пол
	printBday = str(datetime.datetime.strftime(person.birthdate,'%d%m%Y'))
	printBday = printBday[:4]+printBday[-2:]
	for i in range(len(printBday)):
		ws.cell(row=41,column=i+13).value = int(printBday[i])#день рождения
	j=42
	r=0
	for i in range(len(application.abiturient.birthplace)):
		if i>25:
			j=43
			r=25
		ws.cell(row=j,column=i+5-r).value = application.abiturient.birthplace[i]#место рождения
	for i in range (len(passp.docType.value)):
		if i <14:
			ws.cell(row=44,column=i+12).value = passp.docType.value[i]#документ подтверждающий личность
	for i in range(len(str(passp.serialno))):
		ws.cell(row=45,column=i+3).value = int(str(passp.serialno)[i])#серия
	for i in range(len(str(passp.number))):
		ws.cell(row=45,column=i+10).value = int(str(passp.number)[i])#номер
	printPassIssDate = str(datetime.datetime.strftime(passp.issueDate,'%d%m%Y'))
	printPassIssDate = printPassIssDate[:4]+printPassIssDate[-2:]
	for i in range(6):
		ws.cell(row=45,column=i+21).value = int(printPassIssDate[i])#дата выдачи
	j=46
	r=0
	for i in range(len(passp.docIssuer.value)):
		if i!=0:
			if i%24==0:
				j+=1
				r+=23
		ws.cell(row=j,column=i+3-r).value = passp.docIssuer.value[i]#кем выдан
	for i in range(len(application.abiturient.nationality.value)):
		ws.cell(row=49,column=i+5).value = application.abiturient.nationality.value[i]#национальность
	for i in range(len(application.abiturient.citizenship.value)):
		ws.cell(row=50,column=i+5).value = application.abiturient.citizenship.value[i]#гражданство
	for i in range(len(application.abiturient.user.email)):
		ws.cell(row=51,column=i+8).value = application.abiturient.user.email[i]#мыло
	r=8
	for i in range(len(str(snils.serialno))):
		if i != 0:
			if i % 3 == 0:
				r+=1
		ws.cell(row=53,column=i+r).value = str(snils.serialno)[i]#снилс
	###################page2
	if education.doc.isCopy == True:
		ws.cell(row=59,column=7).value = u'Копия'#dokement ob obrazovanii
	else:
		ws.cell(row=59,column=7).value = u'Оригинал'#dokement ob obrazovanii

	Today=datetime.datetime.strftime(datetime.datetime.today(),"%d%m%Y") #tekyshaya data
	Today = Today[:4]+Today[-2:]
	for i in range(6):
		ws.cell(row=59,column=i+21).value = int(Today[i])#Дата
	
	for i in range(len(education.doc.docType.value)):
		ws.cell(row=61,column=i+3).value = education.doc.docType.value[i]#tip dokymenta
	for i in range(len(str(education.doc.serialno))):
		ws.cell(row=61,column=i+13).value = int(str(education.doc.serialno)[i])#lolo
	for i in range(len(str(education.doc.number))):
		ws.cell(row=62,column=i+13).value = int(str(education.doc.number)[i])#lolo
	if education.level.value == 'СОО':
		ws.cell(row=64,column=9).value = 'V'
	if education.level.value == 'НПО':
		ws.cell(row=64,column=12).value = 'V'
	if education.level.value == 'СПО':
		ws.cell(row=64,column=15).value = 'V'
	if education.level.value == 'ВО':
		ws.cell(row=64,column=18).value = 'V'
	eduEnterDate = str(datetime.datetime.strftime(education.enterDate,'%d%m%Y'))
	for i in range(len(eduEnterDate)):
		ws.cell(row=66,column=i+5).value = int(eduEnterDate[i])
	eduLeaveDate = str(datetime.datetime.strftime(education.doc.issueDate, '%d%m%Y'))
	for i in range(len(eduLeaveDate)):
		ws.cell(row=66,column=i+19).value = int(eduLeaveDate[i])
	if len(education.doc.docIssuer.value)>70:
		eduName1 = education.doc.docIssuer.value[:70]
		ws.cell(row=68,column=8).value = eduName1
		eduName2 = education.doc.docIssuer.value[70:]
		ws.cell(row=69,column=8).value = eduName2
	else:
		ws.cell(row=68,column=8).value = education.doc.docIssuer.value#Nazvanie y4ebnogo zavedeniya
	#army
	if hasattr(application.abiturient, 'milit'):
		if application.abiturient.milit.liableForMilit:
			ws.cell(row=102,column=10).value = 'V'
		else:
			ws.cell(row=102,column=18).value = 'V'
		if application.abiturient.milit.isServed:
			ws.cell(row=103,column=7).value = 'V'
			if application.abiturient.milit.yearDismissial is not None:
				for i in range(4):
					ws.cell(row=103,column=i+15).value = int(str(application.abiturient.milit.yearDismissial)[i])
				for i in range(len(application.abiturient.milit.rank.value)):
					if i >14:
						break
					ws.cell(row=104,column=i+6).value = application.abiturient.milit.rank.value[i]
	if application.abiturient.hostel:
		ws.cell(row=106,column=10).value = 'V'
	else:
		ws.cell(row=106, column=14).value = 'V'#obshaga
	for i in range(len(application.abiturient.foreign_lang.value)):
		ws.cell(row=108,column=i+10).value = application.abiturient.foreign_lang.value[i]#inYaz

	for i in range(6):
		ws.cell(row=183,column=i+7).value = int(Today[i])

	###############fiksiki#####################
	thin = Side(border_style="thin", color="000000")
	borderOutside = Border(top=thin, left=thin, right=thin, bottom=thin)
	style_range(ws, 'B12:C13', border=borderOutside)
	style_range(ws, 'B14:C19', border=borderOutside)
	style_range(ws, 'D22:P22', border=borderOutside)
	style_range(ws, 'Q22:V22', border=borderOutside)
	style_range(ws, 'W22:Z22', border=borderOutside)
	style_range(ws, 'Q24:V24', border=borderOutside)
	style_range(ws, 'Q26:V26', border=borderOutside)
	style_range(ws, 'D29:P29', border=borderOutside)
	style_range(ws, 'Q29:Z29', border=borderOutside)
	style_range(ws, 'B117:I117', border=borderOutside)
	style_range(ws, 'J117:T117', border=borderOutside)
	style_range(ws, 'U117:Z117', border=borderOutside)
	style_range(ws, 'N182:Z183', border=borderOutside)
		
	return wb

	""" AMEN """


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill