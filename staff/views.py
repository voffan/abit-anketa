
from django.shortcuts import render, render_to_response, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib import auth
from django.core.context_processors import csrf
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic import ListView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django import template

from staff.applicationPrint import ApplicationPrint
from openpyxl.writer.excel import save_virtual_workbook

import json
from datetime import date
import datetime
from staff.models import Employee, Position, Contacts as ContactsStaff
from anketa.models import ApplicationProfiles, DocImages, EduOrg, ProfileAttrs, EduForm, Attribute, AttrType, Relation, Person, Application, Abiturient, Docs, AttrValue, Profile, Contacts, Address, Education_Prog,  Privilegies, Exams, DepAchieves, Milit, DocAttr, Achievements
from django.contrib.auth.models import User

# Create your views here.
register = template.Library()

def CheckUserIsStaff(user):
	if user.is_staff:
		return True
	else:
		return False
def CheckUserIsSuperuser(user):
	if user.is_superuser:
		return True
	else:
		return False

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def index(request):
	n = 'Anufriev'
	if request.method == 'POST':
		n = request.POST['input1']
	return render(request,'staff\staff_index.html',{'data':{'username':'nik'}}.update(csrf(request)))

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def login(request):
		args = {}
		args.update(csrf(request))
		if request.POST:
			username = request.POST.get('username', '')
			password = request.POST.get('password', '')
			user = auth.authenticate(username=username, password=password)
			if user is not None:
				auth.login(request, user)
				return redirect('/staff/application_list/') # Was '/staff/news'
			else:
					args['login_error'] = "Пользователь не найден"
					return render_to_response('staff\staff_index.html', args)



@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def logout(request):
	auth.logout(request)
	return redirect('/auth')

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Employee_list(request):
	usr = request.user
	if request.method == 'POST':
		ids = request.POST.getlist('selected')
		if 'Delete' in request.POST:
			if usr.is_superuser==1:		
				Del_Employee(ids)
				return HttpResponseRedirect(reverse('staff:employee_list'))
		elif 'Add' in request.POST:
			return HttpResponseRedirect(reverse('staff:employee_add'))
		elif 'Fired' in request.POST:
			if usr.is_superuser==1:
				Fired_Employee(ids)
				return HttpResponseRedirect(reverse('staff:employee_list'))	

	employee_manage = Employee.objects.all()
	#if usr.is_superuser==0: #фильтр видимости для простых сотрудников
	#	empl = Employee.objects.get(user=usr)
	#	employee_manage = employee_manage.filter(department = empl.department)
	if 'finde' in request.POST and len(request.POST['fio'])>0:
		employee_manage = employee_manage.filter(fullname__icontains=request.POST['fio'])

	empl_usr=[]
	for item in employee_manage:
		user = User.objects.select_related('Employee').filter(employee=item.id).first()
		empl_usr.append({'empl':item,'usr':user})
	data={}
	data['employee'] = empl_usr
	context = {'data':data}
	context.update(csrf(request))
	return render(request,'staff\employee_manage.html',  context)

@transaction.atomic
def Fired_Employee(values):	
	if len(values)>0:
		for item in values:
			usr = User.objects.select_related('Employee').filter(employee=item).first()
			if usr.is_staff == True and usr.is_superuser == False:
				if usr.is_active == 0:
					usr.is_active = 1
					usr.save()
				else:
					usr.is_active = 0
					usr.save()

@transaction.atomic
def Del_Employee(values):
	if len(values)>0:
		for item in values:
			usr = User.objects.select_related('Employee').filter(employee=item).first()
			if usr.is_staff == True and usr.is_superuser == False:
				empl = Employee.objects.filter(id = item)
				empl.delete()
				usr.delete() 

#@login_required(login_url = '/auth')
#@user_passes_test(CheckUserIsSuperuser, login_url = '/auth')
@transaction.atomic
def Add_Employee(values):
	empl_id = values.get('user-id','')	
	if len(empl_id)>0:
		employee = Employee.objects.get(pk=empl_id)
		user = employee.user
		user.email = values.get('email','')
		user.save()
	else:
		user = User.objects.create_user(values['username'], values['email'],values['password'])
		user.is_staff = True
		user.save()
		employee = Employee()
		employee.user = user
	dep = EduOrg.objects.get(pk=values['department'])
	posit = Position.objects.get(pk=values['position'])
	employee.department = dep
	employee.first_name = values.get('fname','')
	employee.last_name = values.get('lname','')
	employee.mid_name = values.get('mname','')
	employee.uniemployee = 0
	employee.position = posit
	employee.save()

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsSuperuser, login_url = '/auth')
def AddEmployee(request):
	if request.method == 'POST':
		try:
			Add_Employee(request.POST)
		except Exception as e:
			raise e
		return HttpResponseRedirect(reverse('staff:employee_list'))

	positions = Position.objects.all()
	departments = EduOrg.objects.all()
	Data={}
	Data['departments'] = departments
	Data['positions'] = positions
	context = {'data':Data}
	context.update(csrf(request))
	return render(request,'staff\employee_add.html',context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def EditEmployee(request, employee_id):
	departments = EduOrg.objects.all()
	employee = Employee.objects.get(pk=employee_id)
	positions = Position.objects.all()
	Data={}
	Data['departments'] = departments
	Data['employee']=employee
	Data['positions']=positions
	context = {'data':Data}
	context.update(csrf(request))
	return render(request,'staff\employee_add.html',context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def AddContact(employee, contacts):
	for item in contacts:
		contact = ContactsStaff()
		contact.employee = employee
		contact.contact_type = AttrValue.objects.get(pk=item['id'])
		contact.value = item['value']
		contact.save()

@transaction.atomic
def save_user_profile(user, values):
	user.email = values.get('email','')
	password2 = values['confirm']
	password1 = values['password']
	if password1 == password2 and len(password1) > 0:
		user.set_password(values['password'])
	elif len(password1) > 0:
		raise Exception(u'Пароль и подтверждение не совпадают!!')
	user.save()  
	employee = user.employee_set.get() 
	employee.first_name = values.get('fname','')
	employee.last_name = values.get('lname','')
	employee.mid_name = values.get('mname','')
	if len(values['contacts'])>0:	
		AddContact(employee,[{'id':AttrValue.objects.get(attribute__name__icontains=u'контакт', value__icontains=values.get('contacts_type')).id,'value':values.get('contacts','')}])
	employee.save()

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Employee_Useraccount(request):
	user = request.user
	contacts = user.employee_set.get().contacts_set.all()
	error_message = ''
	Data={}	
	if 'save' in request.POST:
		try:
			#сохранить старые значения user и employee
			#sid = transaction.savepoint()
			save_user_profile(request.user, request.POST)
			Data['success_message'] = u'Успешно изменено'
		except Exception as e:
			error_message = str(e)
			transaction.rollback()
			#восстановить старые значения
		
	Data['contacts']=contacts
	Data['contact_type']=AttrValue.objects.filter(attribute__name__icontains=u'контакт')
	Data['employee']=user.employee_set.get() 
	Data['user']=user
	if len(error_message) > 0:
		Data['error_message'] = error_message

	context = {'data':Data}
	context.update(csrf(request))
	return render(request,'staff\employee_acc.html',context)


@transaction.atomic
def save_exams(exam_data):
	exams = Exams.objects.filter(pk__in=list(exam_data.keys()))	
	for exam in exams:
		exam.points = int(exam_data[str(exam.id)])
		exam.save()
		

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Exam_list(request):
	data={}
	filters={}
	subjects = AttrValue.objects.filter(attribute__name=u'Дисциплина')#spisok disciplin
	##########vse vstypitel'nie ekzameni dlya poly4eniya vsex godov sda4i############
	years=[]
	exams = Exams.objects.filter(exam_examType__value=u'Вступительный')
	for exam in exams:
		years.append(exam.year)
	########################################################################	
	year=datetime.datetime.strftime(datetime.datetime.today(),"%Y") #tekyshiy god
	filters['years'] = int(year)
	if 'years' in request.POST and int(request.POST['years'])>0:
		year = request.POST['years']
		exams = exams.filter(year=request.POST['years'])
		filters['years'] = int(request.POST['years'])

	exams = Exams.objects.filter(exam_examType__value=u'Вступительный', year=year)#spisok ekzamenov na tekyshiy god	

	if 'apply' in request.POST:

		if 'subject' in request.POST and int(request.POST['subject'])>0:
			exams = exams.filter(exam_subjects__id=request.POST['subject'])
			filters['subject_type']=int(request.POST['subject'])
	
		if 'fio' in request.POST and len(request.POST['fio'])>0:
			exams = exams.filter(abiturient__fullname__icontains=request.POST['fio'])
			filters['fio'] = request.POST['fio']

	if 'cancel' in request.POST:
		return HttpResponseRedirect(reverse('staff:exam_list'))

	if 'save' in request.POST:
		exams_id = request.POST.getlist('exam_id')
		exams_points = request.POST.getlist('points')
		save_exams(dict(zip(exams_id, exams_points)))

	data['subjects'] = subjects
	data['exams'] = exams
	data['filters'] = filters
	data['years'] = list(set(years))
	context = {'data':data}
	context.update(csrf(request))
	return render(request, 'staff\Exam_list.html', context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Enlistment(request):
	Data={}
	abiturients = Abiturient.objects.all()
	applications = Application.objects.all()
	Data['abitura'] = abiturients
	Data['applications'] = applications

	context = {'data':Data}
	context.update(csrf(request))
	return render(request,'staff\enlistment.html',context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Application_list (request):
	employee = request.user.employee_set.get()
	if request.user.is_superuser:
		applications = Application.objects.all()
	else:		
		applications = Application.objects.filter(department=employee.department)
	appProfile = ApplicationProfiles.objects.all()
	#applications = Application.objects.select_related('Abiturient').filter(department__id = employee.department.id)
	profiles = Profile.objects.all()
	select = '0'
	selectform = '1'
	selectnapr = '0'
	selectdoc = '0'
	selectcopy = '0'
	selectNumb = '0'
	fname = '0'
	bal1 = '0'
	bal2 = '0'
	dategt = '2016-01-01'
	datelt = '0'
	selectprof = '0'
	selectBplace = '0'
	error_message=''     
	filters={'apply':''}

	if 'apply' in request.GET:

		if 'doctype' in request.GET and int(request.GET['doctype'])>0:
				selectdoc = request.GET['doctype']
				docs = Docs.objects.select_related('Abiturient').filter(docType__id=selectdoc)
				abiturients = [item.abiturient.id for item in docs]
				applications = Application.objects.filter(abiturient__id__in=abiturients)
				filters['doctype']=int(selectdoc)

		if 'iscopy' in request.GET:
			if request.GET['iscopy'] =='1':
				selectcopy = '1'
				docs = Docs.objects.select_related('Abiturient').filter(isCopy=0)
				abiturients = [item.abiturient.id for item in docs]
				applications = applications.filter(abiturient__id__in=abiturients) 
				filters['iscopy'] = selectcopy

			elif request.GET['iscopy'] =='2':
				selectcopy = '2'
				docs = Docs.objects.select_related('Abiturient').filter(isCopy=1)
				abiturients = [item.abiturient.id for item in docs]
				applications = applications.filter(abiturient__id__in=abiturients)
				filters['iscopy'] = selectcopy

		if 'status' in request.GET and int(request.GET['status'])>0:
			select = request.GET['status']
			applications = applications.filter(appState__id=select)
			filters['status']= int(select)

		if 'fio' in request.GET and len(request.GET['fio'])>0:
			fname = request.GET['fio']
			applications=applications.filter(abiturient__fullname__icontains=fname)
			filters['fio'] = fname

		if 'balli1' in request.GET and len(request.GET['balli1'])>0:
			bal1 = int(request.GET['balli1'])
			applications = applications.filter(points__gt=bal1)
			filters['balli1'] = bal1

		if 'balli2' in request.GET and len(request.GET['balli2'])>0:
			bal2 = int(request.GET['balli2'])
			applications = applications.filter(points__lt=bal2)
			filters['balli2'] = bal2

		if 'datedoc1' in request.GET and len(request.GET['datedoc1'])>0:
			dategt = request.GET['datedoc1']
			applications = applications.filter(date__gt=dategt)
			filters['datedoc1'] = dategt

		if 'datedoc2' in request.GET and len(request.GET['datedoc2'])>0:
			datelt = request.GET['datedoc2']
			applications = applications.filter(date__lt=datelt)
			filters['datedoc2'] = datelt

		if 'napravlenie' in request.GET and int(request.GET['napravlenie'])>0:
			selectnapr = request.GET['napravlenie']
			applications = applications.filter(applicationprofiles__profile__profile__edu_prog__id=selectnapr)
			filters['napravlenie'] = int(selectnapr)

		if 'profil' in request.GET and int(request.GET['profil'])>0:
			selectprof = request.GET['profil']
			applications = applications.filter(applicationprofiles__profile__profile__id=selectprof)
			if 'forma' not in request.GET:
				appProfile = appProfile.filter(profile__profile__id = selectprof)
			else:
				applications = applications.filter(applicationprofiles__profile__profile__id=selectprof, applicationprofiles__profile__eduform=EduForm[int(request.GET['forma'])-2][0])
				appProfile = appProfile.filter(profile__profile__id=selectprof, profile__eduform=EduForm[int(request.GET['forma'])-2][0])
				filters['forma'] = selectform
			filters['profil'] = int(selectprof)

		'''if 'forma' in request.GET:
			if request.GET['forma'] =='2':
				applications = applications.filter(applicationprofiles__profile__eduform=u'О')
				selectform = '2'
				appProfile = appProfile.filter(profile__eduform = u'О')
				filters['forma'] = selectform
			if request.GET['forma'] =='3':
				applications = applications.filter(applicationprofiles__profile__eduform=u'З')
				selectform = '3'
				appProfile = appProfile.filter(profile__eduform = u'З')
				filters['forma'] = selectform
			if request.GET['forma'] =='4':
				applications = applications.filter(applicationprofiles__profile__eduform=u'ОЗ')
				selectform = '4'
				appProfile = appProfile.filter(profile__eduform = u'ОЗ')
				filters['forma'] = selectform'''
	
		if 'appNumb' in request.GET and len(request.GET['appNumb'])>0:
			selectNumb = request.GET['appNumb']			
			applications = applications.filter(id=selectNumb)
			if not applications:
				error_message = "vibrannoe zayavlenie vne vashei iyrezdikcii" #ne ny po4ti rabotaet
			filters['appNumb'] = (selectNumb)
		
		if 'birthplace' in request.GET and len(request.GET['birthplace'])>0:
			selectBplace = request.GET['birthplace']
			applications = applications.filter(abiturient__birthplace__icontains=selectBplace)
			filters['birthplace'] = (selectBplace)

	if 'cancel' in request.GET:
		return HttpResponseRedirect(reverse('staff:application_list'))
	
	
	abiturients = [app.abiturient.id for app in applications]
	docs = Docs.objects.select_related('AttrValue').filter(abiturient__id__in = abiturients, docType__value__icontains=u'аттестат')|Docs.objects.select_related('AttrValue').filter(abiturient__id__in = abiturients, docType__value__icontains=u'Диплом')
	apps_with_docs=[]
	for app in applications:
		doc = docs.filter(abiturient__id = app.abiturient.id).first()
		prof = appProfile.filter(application__id=app.id)
		apps_with_docs.append({'app':app, 'doc':doc, 'prof':prof})	
	doctyps = AttrValue.objects.filter(
		attribute__name__icontains=u'об образовании'
	).filter(
		value__icontains=u'Диплом'
	)|AttrValue.objects.filter(
		value__icontains=u'Аттестат'
	)
	profill = AttrValue.objects.filter(attribute__name__icontains=u'Квалификация')
	data={}
	data['errors'] = error_message
	data['applications'] = apps_with_docs
	data['docType'] = doctyps
	data['profill'] = profiles
	data['edu_prog'] = Education_Prog.objects.all()
	data['Docs'] = Docs.objects.all()
	data['Application'] = AttrValue.objects.filter(attribute__name__icontains=u'статус за')
	#data['pages'] = current_page    
	data['filters'] = filters    
	context = {'data':data}
	context.update(csrf(request))
	return render(request,'staff\\application_list.html',context)

################################## backgrid collection json ######################################
@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Backgrid_collection(request):
	employee = request.user.employee_set.get()
	if request.user.is_superuser:
		applications = Application.objects.all()
	else:		
		applications = Application.objects.filter(department=employee.department)
	appProfile = ApplicationProfiles.objects.all()
	applications = Application.objects.select_related('Abiturient').filter(department__id = employee.department.id)
	#profiles = Profile.objects.all()	
	abiturients = [app.abiturient.id for app in applications]
	docs = Docs.objects.select_related('AttrValue').filter(abiturient__id__in = abiturients, docType__value__icontains=u'аттестат')|Docs.objects.select_related('AttrValue').filter(abiturient__id__in = abiturients, docType__value__icontains=u'Диплом')
	apps_with_docs=[]
	result = []
	preress=[]
	for app in applications:
		doc = docs.filter(abiturient__id = app.abiturient.id).first()
		prof = appProfile.filter(application__id=app.id)
		apps_with_docs.append({'app':app, 'doc':doc, 'prof':prof})
		preress.append({'fullname':app.abiturient.fullname,"edu_form":str(prof[0].profile.eduform),"date":str(app.date),"edu_prog":str(prof[0].profile.profile.name),"edu_prof":str(prof[0].profile.profile.edu_prog.qualification.value),"points":int(app.points),"appState":str(app.appState.value)})
	result.append(preress)		
	return HttpResponse(json.dumps(result), content_type="application/json")    
	
def testjson(request):
	return render(request,'staff\\testjson.json')

@login_required(login_url = '/auth')
#@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Application_review (request, application_id):	
	#return HttpResponseRedirect(reverse('staff:application_list'))
	application = Application.objects.select_related('Abiturient').get(pk=application_id)
	passp = application.abiturient.docs_set.filter(docType__value__icontains=u'Паспорт').first()
	passpIn = []
	if passp is None:
		passp = application.abiturient.docs_set.filter(docType__value__icontains=u'Загран').first()
	if passp is None:
		passp = application.abiturient.docs_set.filter(docType__value__icontains=u'Водит').first()
	if passp is None:
		passp = application.abiturient.docs_set.filter(docType__value__icontains=u'Военн').first()
	if passp is not None:
		passpImg = DocImages.objects.filter(doc__id=passp.id)		
		if len(passpImg)>0:
			for img in passpImg:
				passpIn.append({'id':passp.id,'doc':passp,'img':{'id':img.id,'pic':img.image}})
		else:
			passpIn.append({'id':passp.id,'doc':passp})
	Data={}
	
	education = []
	if application.abiturient.education_set.filter().first() is not None:
		education = application.abiturient.education_set.get()
		eduImg = DocImages.objects.filter(doc__id=education.doc.id)		
		edudIn = []
		if len(eduImg)>0:
			for img in eduImg:
				edudIn.append({'education':education,'id':education.doc.id,'doc':education.doc,'img':{'id':img.id,'pic':img.image}})
		else:
			edudIn.append({'education':education,'id':education.doc.id,'doc':education.doc})		
		Data['edud'] = edudIn

	adrtype = AttrValue.objects.filter(attribute__name__icontains=u'Тип адреса')
	rank = AttrValue.objects.filter(attribute__name__icontains=u'Воинское звание')
	snils = application.abiturient.docs_set.filter(docType__value__icontains=u'СНИЛС').first()
	if snils is not None:
		snilsImg = DocImages.objects.filter(doc__id=snils.id)
		Data['snilsImg']=snilsImg
		snils1=[]#####ЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫЫnaidesh ly4she sdelaem ly4she
		if len(str(snils.serialno))>10:
			for i in range(len(str(snils.serialno))):
				if i % 3 == 0 and i != 0:
					snils1+='-'
				snils1+=str(snils.serialno)[i]
			Data['snils'] = snils1
		else:
			Data['snils'] = str(snils.serialno)

	foreign_lang = AttrValue.objects.filter(attribute__name__icontains=u'Изучаемый язык')
	docissuer = AttrValue.objects.filter(attribute__name__icontains=u'выдавший ')
	nationality = AttrValue.objects.filter(attribute__name__icontains=u'национальность')	
	doctype = AttrValue.objects.exclude(value__icontains=u'диплом').exclude(value__icontains=u'аттест').exclude(value__icontains=u'СНИЛС').filter(attribute__name__icontains=u'тип документа')
	Data['edudoctype'] = AttrValue.objects.filter(value__icontains=u'диплом')|AttrValue.objects.filter(value__icontains=u'аттестат')
	contactyp = AttrValue.objects.filter(attribute__name__icontains=u'Тип контакта')
	relation = Relation.objects.filter(abiturient__id=application.abiturient.id)
	relcontacts=[]
	for item in relation:
		text = Person.objects.get(pk=item.person.id).fullname
		cont_value = Contacts.objects.filter(person__id=item.person.id).first().value
		cont_type = item.relType
		relcontacts.append({'id':item.id,'text':text, 'cont':cont_value, 'type':{'name':cont_type.value,'id':cont_type.id}})

	application_profile = ApplicationProfiles.objects.filter(application=application_id)

	exam = Exams.objects.filter(abiturient__id=application.abiturient.id)#, exam_examType__value=u'ЕГЭ'

	Data['relation'] = relcontacts
	Data['rel_type'] = AttrValue.objects.filter(attribute__name__icontains=u'тип связи')
	Data['nationality'] = nationality
	Data['docType'] = doctype
	Data['docissuer'] = docissuer
	Data['adrtype'] = adrtype	
	Data['foreign_lang'] = foreign_lang
	Data['passpIn'] = passpIn
	Data['rank'] = rank
	Data['application']=application
	Data['contacts'] = Contacts.objects.filter(person_id=application.abiturient.id)	
	Data['address'] = Address.objects.filter(pk=application_id)
	Data['education_prog'] = application_profile
	Data['exams'] = exam
	Data['exam_type'] = AttrValue.objects.filter(attribute__name__icontains=u'Тип экзамена')
	Data['exam_subjects'] = AttrValue.objects.filter(attribute__name__icontains=u'Дисциплина')
	Data['privilegies'] = Privilegies.objects.filter(pk=application_id)
	Data['depachieves'] = DepAchieves.objects.filter(pk=application_id)
	if hasattr(application.abiturient, 'milit'):
		Data['milit'] = application.abiturient.milit
	Data['docattr'] = DocAttr.objects.filter(pk=application_id)
	Data['achievements'] = Achievements.objects.filter(pk=application_id)	
	Data['contactyp'] = contactyp

	person=Abiturient.objects.filter(person_ptr__id=application.abiturient.id).first()
	contacts_type = AttrValue.objects.filter(attribute__name__icontains=u'Тип контакта')
	if contacts_type is not None:
		Data['contacts_type']=contacts_type
	person_contacts = person.contacts_set.all()
	if person.contacts_set is not None:
		contacts=[]
		for item in person_contacts:
			cont = {}
			cont['type']=item.contact_type
			cont['value']=item.value
			contacts.append(cont)
		Data['contacts1']=contacts
	
	if request.method =='POST':
		if 'upvote' in request.POST:
			application.appState = AttrValue.objects.filter(value=u'Подтвержденный').first()
			application.save()
		if 'unvote' in request.POST:
			application.appState = AttrValue.objects.filter(value=u'Анулированный').first()
			application.save()
		if 'backspace' in request.POST:
			return HttpResponseRedirect(reverse('staff:application_list'))
		if 'print' in request.POST:
			response = HttpResponse(content=save_virtual_workbook(ApplicationPrint(application,education,application_profile,exam,person,passp,snils)),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			response['Content-Disposition'] = 'attachment; filename=Anketa_SVFU_'+str(application.id)+'.xlsx'
			return response
			

	context = {'data':Data}
	context.update(csrf(request))
	return render(request,'staff\\wizardform.html',context)

@transaction.atomic
def attribute_dels(values):
	dels = values.getlist('selected')
	for item in dels:
		attri_bute = Attribute.objects.filter(id=item)
		attri_bute.delete()

@transaction.atomic
def attrvalue_dels(values):
	dels = values.getlist('selected')
	if len(dels)>0:
		for item in dels:
			attr_value = AttrValue.objects.filter(id=item)
			attr_value.delete()

@transaction.atomic
def attribute_add(values):
	if int(values['attr_id'])<0:
		attri_bute = Attribute(name=values['attr_name'],type_id=values['attrtype'])
	else:
		attri_bute = Attribute.objects.get(pk = values['attr_id'])
		attri_bute.name = values['attr_name']
	attri_bute.save()

@transaction.atomic
def attrvalue_add(attribute, values):
	if int(values['attr_value_id'])<0:
		attr_value_add = AttrValue(value=values['attr_value'], attribute_id=attribute.id)
	else:
		attr_value_add = AttrValue.objects.get(pk=values['attr_value_id'])
		attr_value_add.value = values['attr_value']
	attr_value_add.save()

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Catalogs(request):
	user = User.objects.filter(id = request.user.id).first()
	attribute = Attribute.objects.all()
	Data={}
	if request.method == 'POST':

		if 'filter' in request.POST:
			if len(request.POST.get('attribute1'))>0:
				attribute = attribute.filter(id=request.POST['attribute1'])
		if 'reset' in request.POST:
			return HttpResponseRedirect(reverse('staff:catalogs'))

	if 'delete' in request.POST:
		if user.is_superuser:		
			attribute_dels(request.POST)

	if 'save' in request.POST and len(request.POST.get('attr_name',''))>0:
		if user.is_superuser:
			attribute_add(request.POST)
		
	attribute1 = Attribute.objects.all()    
	attrvalue = AttrValue.objects.all()
	attrtype = AttrType.objects.all()
	Data['attribute1'] = attribute1
	Data['attribute'] = attribute
	Data['attrvalue'] = attrvalue
	Data['attrtype'] = attrtype
	context = {'data':Data}
	context.update(csrf(request))	
	return render(request,'staff\catalogs.html', context)

	
@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Catalogs_attrvalue(request, attribute_id):
	user = User.objects.filter(id = request.user.id).first() 
	attribute = Attribute.objects.get(pk=attribute_id)
	attrvalue = AttrValue.objects.filter(attribute__id=attribute_id)
	if 'delete' in request.POST:
		if user.is_superuser:
			attrvalue_dels(request.POST)
	error_message = ''
	if 'save' in request.POST and len(request.POST['attr_value'])>0:
		if user.is_superuser:
			try:
				attrvalue_add(attribute, request.POST)
			except Exception as e:
				error_message = str(e)
	
	Data={}
	Data['attrvalue'] = attrvalue
	Data['attribute'] = attribute
	Data['error_message'] = error_message
	context = {'data':Data}
	context.update(csrf(request))
	return render(request, 'staff\catalogs_attrvalue.html', context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_orgs(request):
	user = User.objects.filter(id = request.user.id).first()
	edu_orgs = EduOrg.objects.all()
	Data={}
	if request.method == 'POST':
		if 'filter' in request.POST:
			pass
		if 'reset' in request.POST:
			pass

	if 'del' in request.POST:
		if user.is_staff:
			education_delete(request.POST, 0)
	if 'save' in request.POST and len(request.POST.get('edu_org_name',''))>0:
		if user.is_staff:
			edu_org_add(request.POST)

	edu_orgs_type = AttrValue.objects.filter(attribute__name__icontains=u'Тип образовательного')
	edu_orgs_head = EduOrg.objects.filter(head=None)

	Data['edu_orgs'] = edu_orgs
	Data['edu_orgs_type'] = edu_orgs_type
	Data['edu_orgs_head'] = edu_orgs_head
	context = {'data':Data}
	context.update(csrf(request))
	return render(request, 'staff\edu_orgs.html', context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_org_progs(request, edu_org_id):
	user = User.objects.filter(id = request.user.id).first()
	edu_org = EduOrg.objects.get(pk=edu_org_id)
	edu_org_progs = Education_Prog.objects.filter(eduorg__id=edu_org_id)
	Data={}
	error_message=''
	if request.method == 'POST':
		if 'filter' in request.POST:
			pass
		if 'reset' in request.POST:
			pass

	if 'del' in request.POST:
		if user.is_staff:
			education_delete(request.POST, 1)
	if 'save' in request.POST and len(request.POST.get('edu_org_prog_name',''))>0:
		if user.is_staff:
			edu_org_prog_add(edu_org, request.POST)

	edu_orgs_type = AttrValue.objects.filter(attribute__name__icontains=u'Тип образовательного')

	Data['edu_org'] = edu_org
	Data['edu_org_progs'] = edu_org_progs
	Data['error_message'] = error_message
	Data['edu_org_dur'] = AttrValue.objects.filter(attribute__name__icontains=u'Срок обучения')
	Data['edu_org_qual'] = AttrValue.objects.filter(attribute__name__icontains=u'Квалификация')
	context = {'data':Data}
	context.update(csrf(request))
	return render(request, 'staff\edu_org_progs.html', context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_org_prog_profs(request, edu_org_prog_id):
	user = User.objects.filter(id = request.user.id).first()
	edu_org_prog = Education_Prog.objects.get(pk=edu_org_prog_id)
	Data={}
	error_message=''
	if request.method == 'POST':
		if 'filter' in request.POST:
			pass
		if 'reset' in request.POST:
			pass

	if 'del' in request.POST:
		if user.is_staff:
			education_delete(request.POST, 2)
	if 'save' in request.POST and len(request.POST.get('edu_org_prog_prof_name',''))>0:
		if user.is_staff:
			edu_org_prog_prof_add(edu_org_prog, request.POST)

	edu_org_prog_profs = Profile.objects.filter(edu_prog_id=edu_org_prog_id)

	Data['edu_org_prog_profs'] = edu_org_prog_profs
	Data['edu_org_prog'] = edu_org_prog
	Data['error_message'] = error_message
	context = {'data':Data}
	context.update(csrf(request))
	return render(request, 'staff\edu_org_prog_profs.html', context)

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_org_prog_prof_attr(request, edu_org_prog_prof_id):
	user = User.objects.filter(id = request.user.id).first()
	edu_org_prog_prof = Profile.objects.get(pk=edu_org_prog_prof_id)
	Data={}
	error_message=''
	if request.method == 'POST':
		if 'filter' in request.POST:
			pass
		if 'reset' in request.POST:
			pass

	if 'del' in request.POST:
		if user.is_staff:
			education_delete(request.POST, 3)
	if 'save' in request.POST and len(request.POST.get('edu_org_prog_prof_attr_freespaces',''))>0 and len(request.POST.get('edu_org_prog_prof_attr_year'))>0 and len(request.POST.get('startDate'))>0 and len(request.POST.get('endDate'))>0:
		if user.is_staff:
			education_adds(edu_org_prog_prof, request.POST)
			
	edu_org_prog_prof_attr = ProfileAttrs.objects.filter(profile_id=edu_org_prog_prof_id)
	eduform = EduForm

	Data['edu_org_prog_prof'] = edu_org_prog_prof
	Data['edu_org_prog_prof_attr'] = edu_org_prog_prof_attr
	Data['error_message'] = error_message
	Data['eduform'] = eduform
	context = {'data':Data}
	context.update(csrf(request))
	return render(request, 'staff\edu_org_prog_prof_attr.html', context)

@transaction.atomic
def education_adds(attr, values):
	eduform = EduForm
	if int(values['edu_org_prog_prof_attr_id'])<0:
		education_adds = ProfileAttrs()
		education_adds.profile = attr
		education_adds.freespaces = int(values['edu_org_prog_prof_attr_freespaces'])
		education_adds.eduform = values['edu_org_prog_prof_attr_form']
		education_adds.year = int(values['edu_org_prog_prof_attr_year'])
		education_adds.startDate = datetime.datetime.strptime(values['startDate'],'%d/%m/%Y').strftime('%Y-%m-%d')
		education_adds.endDate = datetime.datetime.strptime(values['endDate'],'%d/%m/%Y').strftime('%Y-%m-%d')
		if ProfileAttrs.objects.filter(profile=attr, freespaces=int(values['edu_org_prog_prof_attr_freespaces']),eduform=values['edu_org_prog_prof_attr_form'], year = int(values['edu_org_prog_prof_attr_year']), startDate=datetime.datetime.strptime(values['startDate'],'%d/%m/%Y').strftime('%Y-%m-%d'), endDate = datetime.datetime.strptime(values['endDate'],'%d/%m/%Y').strftime('%Y-%m-%d')).first() == None:
			education_adds.save()
	else:
		education_adds = ProfileAttrs.objects.get(pk = values['edu_org_prog_prof_attr_id'])
		education_adds.profile = attr
		education_adds.freespaces = int(values['edu_org_prog_prof_attr_freespaces'])
		education_adds.eduform = values['edu_org_prog_prof_attr_form']
		education_adds.year = int(values['edu_org_prog_prof_attr_year'])
		education_adds.startDate = datetime.datetime.strptime(values['startDate'],'%d/%m/%Y').strftime('%Y-%m-%d')
		education_adds.endDate = datetime.datetime.strptime(values['endDate'],'%d/%m/%Y').strftime('%Y-%m-%d')
		education_adds.save()

@transaction.atomic
def education_delete(values, attr):
	dels = values.getlist('selected')
	if len(dels)>0:
		for item in dels:
			if attr == 3:
				education_delete = ProfileAttrs.objects.filter(id=item)
			if attr == 2:
				education_delete = Profile.objects.filter(id=item)		
			if attr == 1:
				education_delete = Education_Prog.objects.filter(id=item)
			if attr == 0:
				education_delete = EduOrg.objects.filter(id=item)
			education_delete.delete()

@transaction.atomic
def edu_org_prog_prof_add(edu_org_prog, values):
	if int(values['edu_org_prog_prof_id'])<0:
		edu_org_prog_prof = Profile()
		edu_org_prog_prof.name = values['edu_org_prog_prof_name']
		edu_org_prog_prof.edu_prog = edu_org_prog
		if Profile.objects.filter(name=values['edu_org_prog_prof_name'], edu_prog_id=edu_org_prog.id).first() == None:
			edu_org_prog_prof.save()
	else:
		edu_org_prog_prof = Profile.objects.get(pk=values['edu_org_prog_prof_id'])
		edu_org_prog_prof.name = values['edu_org_prog_prof_name']
		edu_org_prog_prof.edu_prog = edu_org_prog
		edu_org_prog_prof.save()

@transaction.atomic
def edu_org_add(values):
	if int(values['edu_org_id'])<0:
		edu_org = EduOrg(name=values['edu_org_name'], orgtype=AttrValue.objects.filter(id=values['edu_org_type']).first(),head=EduOrg.objects.filter(id=values['edu_org_head']).first())
		if EduOrg.objects.filter(name=edu_org.name, orgtype_id=edu_org.orgtype.id, head_id=edu_org.head.id).first() == None:
			edu_org.save()
	else:
		edu_org = EduOrg.objects.get(pk=values['edu_org_id'])
		edu_org.name = values['edu_org_name']
		if values['edu_org_type'] != '-1':
			edu_org.orgtype = AttrValue.objects.filter(id=values['edu_org_type']).first()
		if values['edu_org_head'] != '-1':
			edu_org.head = EduOrg.objects.filter(id=values['edu_org_head']).first()
		if EduOrg.objects.filter(name=edu_org.name, orgtype=edu_org.orgtype, head=edu_org.head).first() == None:
			edu_org.save()

@transaction.atomic
def edu_org_prog_add(edu_org, values):
	if int(values['edu_org_prog_id'])<0:
		edu_org_prog_add = Education_Prog()
		edu_org_prog_add.eduorg=edu_org
		edu_org_prog_add.name=values['edu_org_prog_name']
		edu_org_prog_add.qualification = AttrValue.objects.get(pk=values['edu_org_qual'])
		edu_org_prog_add.duration = AttrValue.objects.get(pk=values['edu_org_duration'])	
		if Education_Prog.objects.filter(eduorg_id=edu_org.id, name=values['edu_org_prog_name'],qualification_id=values['edu_org_qual'],duration_id=values['edu_org_duration']).first() == None:
			edu_org_prog_add.save()
		
	else:
		edu_org_prog_add = Education_Prog.objects.get(pk=values['edu_org_prog_id'])
		edu_org_prog_add.eduorg=edu_org
		edu_org_prog_add.name = values['edu_org_prog_name']
		edu_org_prog_add.duration = AttrValue.objects.get(pk=values['edu_org_duration'])
		edu_org_prog_add.qualification = AttrValue.objects.get(pk=values['edu_org_qual'])
		if Education_Prog.objects.filter(eduorg_id=edu_org.id, name=values['edu_org_prog_name'],qualification_id=values['edu_org_qual'],duration_id=values['edu_org_duration']).first() == None:
			edu_org_prog_add.save()




#=================================================ajax functions==========================================================
@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_orgs_value(request):
	text = request.GET.get('edu_org_id','')
	item = EduOrg.objects.select_related('AttrValue').get(pk= text)
	if item.orgtype:
		org_type = {'id':item.orgtype.id,'name':item.orgtype.value}
		error_message = ''
	else:
		org_type = {}
		error_message = 'warning'
	if item.head:
		org_head = item.head.id
	else:
		org_head=''
	result = [{ 'name':item.name, 'id':item.id,'type':org_type, 'error':error_message, 'head':org_head}]
	return HttpResponse(json.dumps(result), content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_org_progs_get(request):
	text = request.GET.get('query','')
	item = Education_Prog.objects.select_related('EduOrg').get(pk=text)
	if item.duration:
		dur = item.duration.id
	else:
		dur = ''
	result = [{'name':item.name, 'id':item.id, 'eduorg':item.eduorg.name, 'qual':{'id':item.qualification.id,'name':item.qualification.value}, 'dur':dur}]
	return HttpResponse(json.dumps(result),content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_org_prog_profs_get(request):
	text = request.GET.get('query','')
	item = Profile.objects.select_related('Education_Prog').get(pk=text)
	result=[{'name':item.name, 'edu_prog':item.edu_prog.id, 'id':item.id}]
	return HttpResponse(json.dumps(result), content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Edu_org_prog_prof_attr_get(request):
	text = request.GET.get('query','')
	item = ProfileAttrs.objects.select_related('Profile').get(pk=text)
	result=[{'name':item.profile.name, 'freespaces':item.freespaces, 'eduform':item.eduform, 'id':item.id, 'year':item.year, 'startDate':str(item.startDate), 'endDate':str(item.endDate)}]
	return HttpResponse(json.dumps(result), content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Get_Attrs(request):
	result=[]
	attrs=Attribute.objects.all()
	part = request.GET.get('query','')
	if len(part) > 0:
		attrs = attrs.filter(name__icontains = part)
	for item in attrs:
		result.append({'id':item.id, 'text':item.name})
	return HttpResponse(json.dumps(result), content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Get_Attr(request):
	text = request.GET.get('attribute_id','')
	item = Attribute.objects.select_related('AttrType').get(pk= text)
	result = [{ 'name':item.name, 'id':item.id ,'type':{'id':item.type.id,'name':item.type.name}}]
	return HttpResponse(json.dumps(result), content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Get_Attr_val(request):
	text = request.GET.get('query','')
	attr = AttrValue.objects.select_related('Attribute').get(pk=text)
	result = [{'name':attr.value, 'id':attr.id, 'attribut':attr.attribute.name}]
	return HttpResponse(json.dumps(result),content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Contact_dels(request):
	args = request.POST.get('query','-1')
	contact = ContactsStaff.objects.get(pk=args)
	result = [{'name':contact.value, 'id':contact.id, 'result':1, 'error_message':''}]
	try:
		contact.delete()
	except Exception as e:
		result['result']=0
		result['error_message']=str(e)
	
	return HttpResponse(json.dumps(result),content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Wiz_cont_dels(request):
	args = request.POST.get('query','')
	wiz_cont = Contacts.objects.get(pk=args)
	result=[{'name':wiz_cont.value, 'id':wiz_cont.id, 'result':1, 'error_message':''}]
	try:
		wiz_cont.delete()
	except Exception as e:
		result['result'] =0
		result['error_message'] = str(e)
	return HttpResponse(json.dumps(result), content_type="application/json")

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Wiz_cont_apply(request):
	pass
	args1 = request.POST.get('query1','')
	args2 = request.POST.get('query2','')
	args3 = request.POST.get('query3','')
	wiz_cont = Contacts(person=args1, contact_type=args2, value=args3)
	result=[{'result':1, 'error_message':''}]
	try:
		wiz_cont.save()
	except Exception as e:
		result['result'] = 0
		result['error_message'] = str(e)	
	return HttpResponse(json.dumps(result), content_type="application/json")
	#attri_bute = Attribute(name=values['attr_name'],type_id=values['attrtype'])

@login_required(login_url = '/auth')
@user_passes_test(CheckUserIsStaff, login_url = '/auth')
def Add_exam_to_person(request):
	result={'result':"success"}
	if request.method =='POST':
		try:			
			examsList = request.POST.getlist('examId')
			examPoints = request.POST.getlist('points[1]')

			for i in range(0,len(examsList)):				
				exam = Exams.objects.get(pk=examsList[i])				
				exam.points = examPoints[i]
				exam.save()
		
		except Exception as e:
					result['result']=str(e)
		else:
			pass
		finally:
			pass
	
	
	return HttpResponse(json.dumps(result), content_type="application/json")

def AddDataToPerson(request):
	result="success"
	if request.method == 'POST':
		try:
			page=int(request.POST.get('currentPage',''))
			abit=Abiturient.objects.get(id=request.POST.get('wiz_cont_apply_name',''))			
			if page==1: #Личные данные
				abit.sname=request.POST.get('sname','')
				abit.fname=request.POST.get('name','')
				abit.mname=request.POST.get('mname','')
				abit.birthplace=request.POST.get('birthplace','')
				if(len(request.POST.get('birthday','')))>0:
					abit.birthdate=datetime.datetime.strptime(request.POST.get('birthday',''),'%d/%m/%Y').strftime('%Y-%m-%d')
				if(len(request.POST.get('nation','')))>0:
					abit.nationality=AttrValue.objects.get(pk=request.POST.get('nation',''))
				else:
					abit.nationality=None
				if(len(request.POST.get('citizenship','')))>0:
					abit.citizenship=AttrValue.objects.get(pk=request.POST.get('citizenship',''))
				if(len(request.POST.get('sex','')))>0:
					abit.sex=request.POST.get('sex','')
			if page==2:
				doctype = abit.docs_set.filter(docType__attribute__name__icontains=u'удостоверяющего личность').first()
				if doctype is None:
					doctype=Docs()
					doctype.abiturient=abit
				else:
					doctype.number=None
					doctype.serialno=None
					doctype.issueDate=None
				if(len(request.POST.get('doctype',''))>0):
					doctype.docType=AttrValue.objects.get(pk=request.POST.get('doctype',''))
				if(len(request.POST.get('serialdoc',''))>0):
					doctype.serialno=int(request.POST.get('serialdoc',''))
				if(len(request.POST.get('numberdoc',''))>0):
					doctype.number=int(request.POST.get('numberdoc',''))
				if(len(request.POST.get('datedoc',''))>0):
					doctype.issueDate=datetime.datetime.strptime(request.POST.get('datedoc',''),'%d/%m/%Y').strftime('%Y-%m-%d')
				if(len(request.POST.get('docissuer',''))>0):
					doctype.docIssuer=AttrValue.objects.get(pk=request.POST.get('docissuer',''))
				doctype.save()
				edudoc = abit.education_set.first()
				if edudoc is None:
					edudoc=Education()
					edudoc.abiturient=abit
					doc = Docs()
					doc.abiturient=abit
				else:
					doc = edudoc.doc
				if(len(request.POST.get('edudoctype','')))>0:
					doc.docType=AttrValue.objects.get(pk=request.POST.get('edudoctype',''))
				if(len(request.POST.get('serialedudoc',''))>0):
					doc.serialno=int(request.POST.get('serialedudoc',''))
				if(len(request.POST.get('numberedudoc',''))>0):
					doc.number=int(request.POST.get('numberedudoc',''))
				if(len(request.POST.get('dateexiting',''))>0):
					doc.issueDate=datetime.datetime.strptime(request.POST.get('dateexiting',''),'%d/%m/%Y').strftime('%Y-%m-%d')
				if(len(request.POST.get('preveduname','')))>0:
					doc.docIssuer=AttrValue.objects.get(pk=request.POST.get('preveduname',''))
				doc.save()
				edudoc.doc = doc
				datejoining = request.POST.get('datejoining','')
				if len(datejoining)>0:
					edudoc.enterDate=datetime.datetime.strptime(datejoining,'%d/%m/%Y').strftime('%Y-%m-%d')
				if doc.docType.value == "Аттестат":
					edudoc.level = AttrValue.objects.filter(attribute__name__icontains=u'Предыдущее образование').filter(value__icontains=u'СОО').first()
				else:
					prevedu = request.POST.get('prevedu','')
					if len(prevedu)>0:
						if prevedu == "npo":
							edudoc.level = AttrValue.objects.filter(attribute__name__icontains=u'Предыдущее образование').filter(value__icontains=u'НПО').first()
						else:
							if prevedu == "spo":
								edudoc.level = AttrValue.objects.filter(attribute__name__icontains=u'Предыдущее образование').filter(value__icontains=u'СПО').first()
							else:
								edudoc.level = AttrValue.objects.filter(attribute__name__icontains=u'Предыдущее образование').filter(value__icontains=u'ВПО').first()
				edudoc.save()
				
				snils = abit.docs_set.filter(docType__value__icontains=u'CНИЛС').first()
				if snils is None:
					snils = Docs()
					snils.docType=AttrValue.objects.filter(value__icontains=u'СНИЛС').first()
					snils.abiturient=abit
				if (len(request.POST.get('inila',''))>0):
					snils.serialno=int(request.POST.get('inila',''))
				snils.docIssuer=doctype.docIssuer
				snils.save()
			
			if page==3:
				"""
				if request.POST.get('adrstype','')=="perm":
					adrs_type=u'По прописке'
				else:
					adrs_type=u'Фактический'
				adrs=Address.objects.filter(abiturient=abit).filter(adrs_type__value__icontains=adrs_type).first()
				if adrs is None:
					adrs=Address()
					adrs.abiturient=abit
				adrs.adrs_type=AttrValue.objects.filter(value__icontains=adrs_type).first()
				adrs.zipcode=request.POST.get('adrsindex','')
				adrs.street=Street.objects.filter(name__icontains=request.POST.get('street','')).first()
				adrs.house=request.POST.get('adrshouse','')
				adrs.building=request.POST.get('adrsbuilding','')
				adrs.flat=request.POST.get('adrsflat','')
				if ((request.POST.get('adrsisthesame','')) == "yes"):
					adrs.adrs_type_same=True
					#adrs.adrs_type=AttrValue.objects.filter(value__icontains=u'прописке').first()
					Address.objects.filter(abiturient=abiturient).delete()
				else:
					adrs.adrs_type_same=False
				adrs.save()"""
				if abit.contacts_set.all() is not None:
					Contacts.objects.filter(person=abit).delete()
				contacttype = request.POST.getlist('contacttype')
				contactvalue = request.POST.getlist('contactvalue')
				for i in range(0, len(contacttype)):
					if len(contacttype[i])>0 and len(contactvalue[i])>0:
						contact = Contacts()
						contact.person = abit
						contact.contact_type=AttrValue.objects.filter(pk=contacttype[i]).first()
						contact.value = contactvalue[i]
						contact.save()

				if Relation.objects.filter(abiturient = abit) is not None:
					Relation.objects.filter(abiturient = abit).delete()
				reltype = request.POST.getlist('rel_contacttype')
				relcontactvalue = request.POST.getlist('rel_contactvalue')				
				arr = request.POST.getlist('rel_fio')				
				for i in range(0, len(reltype)):
					if len(reltype[i])>0 and len(relcontactvalue[i])>0 and len(arr[i])>0:
						relperson = Person()
						fio =[]
						fio = arr[i].split(' ')
						relperson.sname = fio[0]
						relperson.fname = fio[1]
						relperson.mname = fio[2]
						relperson.fullname = arr[i]
						relperson.sex = 'М'
						relperson.birthdate = "2000-11-11"
						relperson.save()
						contact = Contacts()
						contact.person = relperson
						contact.contact_type = AttrValue.objects.filter(value__icontains=u'телефон').first()
						contact.value = relcontactvalue[i]
						contact.save()
						newrel = Relation()
						newrel.person = relperson
						newrel.abiturient = abit
						newrel.relType = AttrValue.objects.filter(pk=reltype[i]).first()
						newrel.save()	
			if page==4:
				if abit.exams_set.filter(exam_examType__value=u'ЕГЭ') is not None:
					Exams.objects.filter(abiturient=abit).filter(exam_examType__value=u'ЕГЭ').delete()
					#abit.exams_set.all().delete a cho ne rabotaet
				examtype=AttrValue.objects.filter(attribute__name__icontains=u'Тип экзамена').filter(value__icontains=u'ЕГЭ').first()
				exams = request.POST.getlist('egeDisc')
				points = request.POST.getlist('egePoints')
				years=request.POST.getlist('egeYear')
				for i in range(0, len(exams)):
					exam = Exams()
					exam.abiturient = abit
					exam.exam_examType=examtype
					exam.exam_subjects=AttrValue.objects.filter(attribute__name__icontains=u'Дисциплина').filter(pk=exams[i]).first()
					exam.points=points[i]
					exam.year=years[i]
					exam.save()
				if len(request.POST.get('additionalExams','')) > 0:
					add_exams=request.POST.get('additionalExams','').split(',')
					specusl = False
					if request.POST.get('specusl','') == "yes":
						specusl=True
					if abit.exams_set.filter(exam_examType__value=u'Вступительный') is not None:
						Exams.objects.filter(abiturient=abit).filter(exam_examType__value=u'Вступительный').delete()
					for item in add_exams:
						add_exam = Exams()
						add_exam.abiturient=abit
						add_exam.exam_examType = AttrValue.objects.filter(attribute__name__icontains=u'Тип экзамена').filter(value__icontains=u'Вступительный').first()
						add_exam.exam_subjects=AttrValue.objects.filter(attribute__name__icontains=u'Дисциплина').filter(pk=item).first()
						add_exam.year=2016
						add_exam.special = specusl
						add_exam.save()
			"""
			if(page==5):

			if(page==6):
			"""
			if page==7:
				if(len(request.POST.get('hostel','')))>0:
					abit.hostel=False
					if(request.POST.get('hostel','')=="yes"):
						abit.hostel=True
				if(len(request.POST.get('flang','')))>0:
					abit.foreign_lang=AttrValue.objects.get(pk=request.POST.get('flang',''))

				if Milit.objects.filter(abiturient = abit).first() is not None:
					Milit.objects.filter(abiturient = abit).first().delete()
				milit = Milit()
				milit.abiturient=abit
				if (len(request.POST.get('liableForMilit','')))>0:
					if request.POST.get('liableForMilit','')=="yes":
						milit.liableForMilit=True
						if (len(request.POST.get('isServed','')))>0:
							if request.POST.get('isServed','')=="yes":
								milit.isServed=True
								if(len(request.POST.get('rank','')))>0:
									milit.rank=AttrValue.objects.get(pk=request.POST.get('rank'))
								if (len(request.POST.get('yeararmy','')))>0:
									milit.yearDismissial=int(request.POST.get('yeararmy',''))
									# АХАХАХАХАХ ХКАКОЙ ККРАСИВЫЙ КОД АХАХАХАХАХХААХХА
				milit.save()			
			abit.save()			
		except Exception as e:
					result=str(e)
	return HttpResponse(json.dumps(result), content_type="application/json")

	'''wb = Workbook()
	ws = wb.active;
	ws.title = u'Операции магазина'
	ws['A1'] = u'Операции магазина ' + shop.name

	_row = 2
	_column = 1
	if 'filters' in request.session and 'operations' in request.session['filters']:
		if 'op_date' in request.session['filters']['operations']:
			ws.cell(row=_row, column=_column).value = u'На дату: ' + request.session['filters']['operations']['op_date']
			_row += 1
		if 'oper' in request.session['filters']['operations']:
			ws.cell(row=_row, column=_column).value = u'По операции: ' + Operation.objects.get(
				pk=request.session['filters']['operations']['oper']).name
			_row += 1
		if 'order' in request.session['filters']['operations']:
			ws.cell(row=_row, column=_column).value = u'По номеру заказа: ' + request.session['filters']['operations'][
				'order']
			_row += 1
		if 'employee' in request.session['filters']['operations']:
			ws.cell(row=_row, column=_column).value = u'По сотруднику' + ShopStaff.objects.get(
				person__pk=request.session['filters']['operations']['employee']).person.name
			_row += 1
		if 'item' in request.session['filters']['operations']:
			ws.cell(row=_row, column=_column).value = u'' + Items.objects.get(
				pk=request.session['filters']['operations']['item']).name
			_row += 1

	_row += 2
	ws.cell(row=_row, column=1).value = u'№'
	ws.cell(row=_row, column=2).value = u'Дата'
	ws.cell(row=_row, column=3).value = u'Операция'
	ws.cell(row=_row, column=4).value = u'Заказ'
	ws.cell(row=_row, column=5).value = u'Сотрудник'
	ws.cell(row=_row, column=6).value = u'Товар'
	ws.cell(row=_row, column=7).value = u'Кол-во'
	ws.cell(row=_row, column=8).value = u'Стоимость'
	i = 1
	_row += 1
	for operation in operations:
		ws.cell(row=_row, column=1).value = i
		ws.cell(row=_row, column=2).value = operation.operation_date.strftime('%d.%m.%Y')
		ws.cell(row=_row, column=3).value = operation.operation.name
		if operation.order is not None:
			ws.cell(row=_row, column=4).value = operation.order.id
		ws.cell(row=_row, column=5).value = operation.user.person.name
		ws.cell(row=_row, column=6).value = operation.item.name
		ws.cell(row=_row, column=7).value = operation.item_vol
		ws.cell(row=_row, column=8).value = operation.item_sum
		_row += 1
		i += 1

	response = HttpResponse(content=save_virtual_workbook(wb),#save_virtual_workbook находится в openpyxl
							content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=opertions.xlsx'
	return response'''